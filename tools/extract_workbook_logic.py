import json
import re
import sys
from pathlib import Path

import openpyxl


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        match = re.search(r'COMPUTED_VALUE"""\),(.+)\)$', value)
        if match:
            value = match.group(1)
        return value
    return value


def rows_values(sheet, start, end, cols):
    rows = []
    for row_idx in range(start, end + 1):
        row = [sheet.cell(row_idx, col_idx).value for col_idx in range(1, cols + 1)]
        if any(value is not None for value in row):
            rows.append(row)
    return rows


def main() -> None:
    workbook_path = Path(sys.argv[1])
    wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)

    calculator = wb_formula["Calculator"]
    calculator_values = wb_values["Calculator"]
    calculations = wb_formula["Calculations"]
    settings_values = wb_values["Settings"]
    labor_values = wb_values["Labor"]
    vehicles_values = wb_values["Vehicles"]
    warnings_values = wb_values["Warnings"]
    zones_values = wb_values["Zones"]

    item_formula_cells = [
        "G6", "J6", "K6", "L6", "M6", "O6",
        "K38", "G38", "J38", "O38", "M1", "N1", "M4"
    ]
    calculation_formula_cells = [
        "B4", "C4", "D4", "E4", "F4", "G4", "H4", "I4", "J4",
        "BE4", "BH4"
    ]

    payload = {
        "sheets": [
            {"name": ws.title, "rows": ws.max_row, "columns": ws.max_column}
            for ws in wb_formula.worksheets
        ],
        "calculator": {
            "inputs": {
                "pickup_zone": calculator_values["C2"].value,
                "delivery_zone": calculator_values["C3"].value,
                "distance": calculator_values["D4"].value,
                "pickup_team": calculator_values["G2"].value,
                "interstate_team": calculator_values["G3"].value,
                "delivery_team": calculator_values["G4"].value,
                "pickup_vehicle": calculator_values["H2"].value,
                "interstate_vehicle": calculator_values["H3"].value,
                "delivery_vehicle": calculator_values["H4"].value,
            },
            "outputs": {
                "price": calculator_values["M1"].value,
                "price_per_cuft_excluding_storage": calculator_values["N1"].value,
                "insurance": calculator_values["M2"].value,
                "storage": calculator_values["M3"].value,
                "cost_of_service": calculator_values["M4"].value,
                "total_volume": calculator_values["K38"].value,
                "total_weight": calculator_values["G38"].value,
            },
            "formulas": {cell: str(calculator[cell].value) for cell in item_formula_cells},
        },
        "calculations_formulas": {
            cell: str(calculations[cell].value)
            for cell in calculation_formula_cells
            if calculations[cell].value is not None
        },
        "settings": rows_values(settings_values, 1, 62, 3),
        "labor": rows_values(labor_values, 1, 24, 6),
        "vehicles": rows_values(vehicles_values, 1, 12, 18),
        "warnings": rows_values(warnings_values, 1, 7, 3),
        "zones": rows_values(zones_values, 1, 11, 10),
    }

    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
