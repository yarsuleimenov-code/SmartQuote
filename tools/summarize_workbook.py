import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl


def cell_value(cell):
    value = cell.value
    if value is None:
        return ""
    return str(value)


def non_empty_rows(sheet, max_rows=80, max_cols=20):
    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, max_rows), max_col=min(sheet.max_column, max_cols)):
        values = [cell.value for cell in row]
        if any(value is not None for value in values):
            rows.append(values)
    return rows


def formulas(sheet, max_items=120):
    result = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                result.append({"cell": cell.coordinate, "formula": cell.value})
                if len(result) >= max_items:
                    return result
    return result


def main() -> None:
    workbook_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)

    output = {
        "sheets": [
            {"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column}
            for ws in wb.worksheets
        ],
        "sheet_rows": {},
        "formulas": {},
        "calculator_inputs_outputs": {},
        "formula_references": {},
    }

    for name in ["Техзадание", "Settings", "Labor", "Warnings", "Zones", "Checking", "Calculator", "Calculations", "Vehicles", "History"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            output["sheet_rows"][name] = non_empty_rows(sheet)
            output["formulas"][name] = formulas(sheet, 160)

    calc = wb["Calculator"]
    calc_values = wb_values["Calculator"]
    interesting_cells = [
        "C2", "C3", "D4", "G2", "G3", "G4", "H2", "H3", "H4",
        "M1", "N1", "M2", "M3", "M4", "K38", "G38", "J38", "O38"
    ]
    output["calculator_inputs_outputs"] = {
        cell: {"formula_or_value": calc[cell].value, "computed": calc_values[cell].value}
        for cell in interesting_cells
    }

    ref_counter = Counter()
    for sheet in wb.worksheets:
        for item in formulas(sheet, 10_000):
            for ref in ["Settings!", "Labor!", "Vehicles!", "Zones!", "Warnings!", "Checking!", "Calculations!", "Calculator!"]:
                if ref in item["formula"]:
                    ref_counter[f"{sheet.title}->{ref[:-1]}"] += 1
    output["formula_references"] = dict(ref_counter)

    print(json.dumps(output, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
