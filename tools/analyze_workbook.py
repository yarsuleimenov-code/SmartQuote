import json
import sys
from pathlib import Path

import openpyxl


def cell_payload(cell, value_sheet):
    return {
        "cell": cell.coordinate,
        "value": cell.value,
        "computed": value_sheet[cell.coordinate].value,
    }


def main() -> None:
    workbook_path = Path(sys.argv[1])
    formulas_book = openpyxl.load_workbook(workbook_path, data_only=False)
    values_book = openpyxl.load_workbook(workbook_path, data_only=True)

    summary = {}
    for sheet in formulas_book.worksheets:
        values_sheet = values_book[sheet.title]
        samples = []
        formulas = []
        non_empty_count = 0

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                non_empty_count += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if len(formulas) < 80:
                        formulas.append(cell_payload(cell, values_sheet))
                elif len(samples) < 60:
                    samples.append(cell_payload(cell, values_sheet))

        summary[sheet.title] = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "non_empty_count": non_empty_count,
            "formula_count": sum(
                1
                for row in sheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            ),
            "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
            "sample_values": samples,
            "sample_formulas": formulas,
        }

    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
